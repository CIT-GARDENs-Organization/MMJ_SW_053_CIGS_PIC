import os
import glob
import pandas as pd
from datetime import datetime

# ログインユーザー名を取得
username = os.environ.get("USERNAME")

# ベースパスの構築
base_path = fr"C:\Users\{username}\Chiba Institute of Technology\5号機 - ドキュメント\General\00_MMJ文書\ICD\MMJ_ICD_018_IICD"

# ファイル名キーワード
filename_keyword = "Memory_map"

# ファイル検索パターン
search_pattern = os.path.join(base_path, f"*{filename_keyword}*.xlsx")

def generate_memory_map_header(excel_file):
    """Excelファイルからメモリマップヘッダーファイルを生成"""
    try:
        import time
        max_retries = 3
        retry_delay = 1
        
        for attempt in range(max_retries):
            try:
                # openpyxlで読み取り専用モードを試行
                from openpyxl import load_workbook
                print(f"Attempt {attempt + 1}: Opening Excel file in read-only mode...")
                wb = load_workbook(excel_file, read_only=True, data_only=True)
                available_sheets = wb.sheetnames
                print(f"Available sheets: {available_sheets}")
                
                # データを読み込む
                excel_data_success = True
                break
                
            except PermissionError as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    print(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                    continue
                else:
                    raise e
            except Exception as e:
                print(f"Read-only mode failed: {e}")
                # フォールバック: pandasで試行
                xl_file = pd.ExcelFile(excel_file, engine='openpyxl')
                available_sheets = xl_file.sheet_names
                excel_data_success = True
                break
        
        header_content = f"""#ifndef SMF_MEMORY_MAP_H
#define SMF_MEMORY_MAP_H

/*
 * Auto-generated Memory Map Header
 * Generated from: {os.path.basename(excel_file)}
 * Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
 */

// Base Memory Configuration
#define SMF_SIZE 0x08000000 // 128MB
#define SMF_START_ADDRESS 0x00000000
#define SMF_END_ADDRESS   0x07FFFFFF

"""
        
        # メモリマップシートを処理
        memory_map_sheet = None
        for sheet_name in available_sheets:
            if 'メモリマップ' in sheet_name:
                memory_map_sheet = sheet_name
                break
        
        if memory_map_sheet:
            print(f"\n=== Processing Sheet: {memory_map_sheet} ===")
            df = pd.read_excel(excel_file, sheet_name=memory_map_sheet, engine='openpyxl')
            print(f"Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            
            # ヘッダー行を見つける
            header_row = None
            for i in range(min(10, len(df))):
                row_str = ' '.join([str(df.iloc[i][col]) for col in df.columns if pd.notna(df.iloc[i][col])])
                if 'System' in row_str and ('Device' in row_str or 'ID' in row_str):
                    header_row = i
                    print(f"Found header row at index {i}")
                    break
            
            if header_row is not None:
                # B列(System)とC列(Device ID)のデータを抽出
                header_content += f"\n// ===== Device IDs from {memory_map_sheet} =====\n"
                
                # デバッグ情報を追加
                print(f"\nHeader row content:")
                for j, col in enumerate(df.columns):
                    if pd.notna(df.iloc[header_row, j]):
                        print(f"  Column {j}: {df.iloc[header_row, j]}")
                
                print(f"\nProcessing data rows {header_row + 1} to {len(df)}:")
                
                # デバイスIDをExcelデータから動的に抽出
                device_mappings = {}  # {upper_nibble: device_name} の辞書
                
                # DataIDとSystem列からデバイス名とIDのマッピングを抽出
                for i in range(header_row + 1, len(df)):
                    try:
                        # DataID列（D列、インデックス3）からデバイスIDを取得
                        if len(df.columns) > 3 and pd.notna(df.iloc[i, 3]):
                            data_id_raw = str(df.iloc[i, 3]).strip()
                            if data_id_raw and data_id_raw != 'nan' and data_id_raw != '-':
                                # DataIDの上位ニブル（0x1, 0x2, 0xC など）を抽出
                                if len(data_id_raw) >= 3:  # "0xX" の最小長
                                    upper_nibble = data_id_raw[:3]  # "0x1", "0x2", "0xC" など
                                    
                                    # System列（B列、インデックス1）とAssignment列（C列、インデックス2）からデバイス名を取得
                                    device_name = None
                                    
                                    # まずAssignment (Device ID)列（C列、インデックス2）を確認
                                    if len(df.columns) > 2 and pd.notna(df.iloc[i, 2]):
                                        assignment_name = str(df.iloc[i, 2]).strip().upper()
                                        if assignment_name and assignment_name != 'NAN':
                                            # "MAIN PIC\n(1)" や "COM PIC\n(2)" などから "MAIN", "COM" を抽出
                                            # 改行、括弧、数字を除去してクリーンアップ
                                            device_name = assignment_name.replace(" PIC", "").replace(" MCU", "")
                                            device_name = device_name.split('\n')[0].strip()  # 改行で分割して最初の部分のみ
                                            device_name = ''.join(c for c in device_name if c.isalpha())  # アルファベットのみ
                                            print(f"      Debug: Row {i}, Assignment={assignment_name} -> {device_name}")
                                    
                                    # Assignment列が空の場合、System列（B列、インデックス1）を確認
                                    if not device_name and len(df.columns) > 1 and pd.notna(df.iloc[i, 1]):
                                        system_name = str(df.iloc[i, 1]).strip().upper()
                                        if system_name and system_name != 'NAN':
                                            # "MAIN PIC", "COM PIC" などから "MAIN", "COM" を抽出
                                            device_name = system_name.replace(" PIC", "").replace(" MCU", "")
                                            device_name = device_name.split('\n')[0].strip()  # 改行で分割して最初の部分のみ
                                            device_name = ''.join(c for c in device_name if c.isalpha())  # アルファベットのみ
                                            print(f"      Debug: Row {i}, System={system_name} -> {device_name}")
                                    
                                    # デバイス名が取得できた場合のみマッピングに追加
                                    if device_name and upper_nibble not in device_mappings:
                                        device_mappings[upper_nibble] = device_name
                                        print(f"      Debug: Added mapping {upper_nibble} -> {device_name}")
                    except Exception:
                        continue
                
                # 抽出されたデバイスIDを出力
                for upper_nibble, device_name in sorted(device_mappings.items()):
                    device_id = upper_nibble + "0"  # "0x1" -> "0x10"
                    header_content += f"#define {device_name}_DEVICE_ID {device_id}\n"
                    print(f"    Added: {device_name}_DEVICE_ID -> {device_id}")
                
                # メモリアドレスマッピングを追加
                header_content += f"\n// ===== Memory Address Mappings from {memory_map_sheet} =====\n"
                
                # 重複チェック用のセット
                generated_definitions = set()
                
                # Excelデータから"Sub Assignment", "Start address", "End address"を抽出
                current_device_id = None  # 結合セル対応のため、現在のデバイスIDを保持
                
                for i in range(header_row + 1, len(df)):
                    try:
                        sub_assignment = None
                        start_address = None
                        end_address = None
                        device_id = None
                        
                        # デバッグ: 行の内容を確認
                        row_values = []
                        for j in range(min(10, len(df.columns))):
                            if pd.notna(df.iloc[i, j]):
                                val = str(df.iloc[i, j]).strip()
                                if val:
                                    row_values.append(f"Col{j}:{val}")
                        
                        # Sub Assignment列（E列、インデックス4）
                        if len(df.columns) > 4 and pd.notna(df.iloc[i, 4]):
                            sub_assignment = str(df.iloc[i, 4]).strip()
                        
                        # Start address列（I列、インデックス8）
                        if len(df.columns) > 8 and pd.notna(df.iloc[i, 8]):
                            start_address = str(df.iloc[i, 8]).strip()
                        
                        # End address列（J列、インデックス9）
                        if len(df.columns) > 9 and pd.notna(df.iloc[i, 9]):
                            end_address = str(df.iloc[i, 9]).strip()
                        
                        # DataID列（D列、インデックス3）からデバイスIDを取得
                        data_id = None
                        if len(df.columns) > 3 and pd.notna(df.iloc[i, 3]):
                            data_id_raw = str(df.iloc[i, 3]).strip()
                            if data_id_raw and data_id_raw != 'nan' and data_id_raw != '-':
                                data_id = data_id_raw
                        
                        # デバッグ情報を表示（CIGS関連のみ）
                        if i < header_row + 90 and (data_id and "0xC" in data_id):
                            print(f"    Row {i}: DataID={data_id}, Sub={sub_assignment}, Start={start_address}, End={end_address}")
                            print(f"      *** CIGS RELATED ***")
                        
                        # データが有効な場合のみ処理
                        if sub_assignment and start_address and end_address and data_id:
                            # DataIDに基づいて定義名を生成（汎用パターンマッチング）
                            definition_name = None
                            sub_clean = sub_assignment.upper().strip()
                            
                            # Sub AssignmentからDevice IDを除去（デバイス名の重複のみ）
                            device_id_patterns = [
                                " (1)", " (2)", " (6)", " (7)", " (8)", " (9)", " (B)", " (C)",
                                "(1) ", "(2) ", "(6) ", "(7) ", "(8) ", "(9) ", "(B) ", "(C) ",
                                "MAIN ", "COM ", "APRS ", "CAM ", "CHO ", "SATO ", "BHU ", "CIGS "
                            ]
                            for pattern in device_id_patterns:
                                sub_clean = sub_clean.replace(pattern, " ").strip()
                            
                            # デバイスプレフィックスをDataIDから決定（動的）
                            device_prefix = None
                            for upper_nibble, device_name in device_mappings.items():
                                if data_id.startswith(upper_nibble):  # "0x1", "0x2", "0xC" などの比較
                                    device_prefix = device_name
                                    break
                            
                            if not device_prefix:
                                continue
                            
                            # Sub Assignment文字列をそのままdefine名に変換
                            # スペース、ハイフン、アンパサンドをアンダースコアに変換
                            safe_name = sub_clean.replace(' ', '_').replace('-', '_').replace('&', '_')
                            safe_name = safe_name.replace('(', '').replace(')', '').replace(',', '_')
                            
                            # 不要な文字を除去し、連続するアンダースコアを1つにまとめる
                            safe_name = ''.join(c for c in safe_name if c.isalnum() or c == '_')
                            while '__' in safe_name:
                                safe_name = safe_name.replace('__', '_')
                            safe_name = safe_name.strip('_')
                            
                            # 最終的な定義名を作成
                            if safe_name:
                                definition_name = f"{device_prefix}_{safe_name}"
                            else:
                                # 空の場合はスキップ
                                continue
                            
                            # アドレスを16進数形式に正規化
                            if not start_address.startswith('0x') and start_address.isalnum():
                                start_address = f"0x{start_address.upper()}"
                            if not end_address.startswith('0x') and end_address.isalnum():
                                end_address = f"0x{end_address.upper()}"
                            
                            # 重複チェック - 同じ定義名が既に存在する場合はスキップ
                            start_def = f"{definition_name}_START_ADDRESS"
                            end_def = f"{definition_name}_END_ADDRESS"
                            
                            if start_def not in generated_definitions:
                                # 定義を追加
                                header_content += f"#define {start_def} {start_address}\n"
                                header_content += f"#define {end_def} {end_address}\n"
                                generated_definitions.add(start_def)
                                generated_definitions.add(end_def)
                                print(f"    Added: {definition_name} -> {start_address} to {end_address}")
                            else:
                                print(f"    Skipped duplicate: {definition_name}")
                    except Exception as e:
                        continue
            else:
                print("❌ Header row not found in memory map sheet")
        else:
            print("❌ Memory map sheet not found")
        
        header_content += "\n#endif // SMF_MEMORY_MAP_H\n"
        
        print("\n=== Final Generated Header Content ===")
        print(header_content)
        
        return header_content
        
    except Exception as e:
        print(f"❌ Excelファイル読み込みエラー: {e}")
        print("❌ Excelファイルにアクセスできません。ファイルが開いていないか確認してください。")
        print("   - Excelファイルを閉じてから再実行してください")
        print("   - ファイルの権限を確認してください")
        print("   - ネットワーク接続を確認してください")
        return None

# 実行と表示
print(f"Search Pattern: {search_pattern}")
matches = glob.glob(search_pattern)
if matches:
    excel_file = matches[0]
    print(f"✅ ファイル見つかった: {excel_file}")
    
    # ヘッダーファイル生成を試行
    header_content = generate_memory_map_header(excel_file)
    
    if header_content:
        # ファイルに保存（絶対パスを使用）
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(script_dir)
        output_file = os.path.join(project_root, "lib", "tool", "mmj_smf_memorymap.h")

        # ディレクトリが存在しない場合は作成
        output_dir = os.path.dirname(output_file)
        os.makedirs(output_dir, exist_ok=True)
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(header_content)
            print(f"\n✅ ヘッダーファイルが生成されました: {output_file}")
        except Exception as e:
            print(f"❌ ファイル保存エラー: {e}")
    else:
        print("❌ ヘッダーファイルの生成に失敗しました")
        print("   処理を中断します")
        exit(1)
    
else:
    print("❌ Excelファイルが見つかりませんでした")
    print(f"   検索パス: {search_pattern}")
    print("   以下を確認してください:")
    print("   - ファイルが正しい場所にあるか")
    print("   - ファイル名にMemory_mapが含まれているか")
    print("   - ネットワークドライブが接続されているか")
    exit(1)
